import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from item import Item
from population import Population

class GeneticAlgorithm:
    def __init__(self, filename, population_size, crossover_rate, mutation_rate, num_generations):
        self.items, self.capacity = self.load_items_from_file(filename)
        self.population_size = population_size
        self.crossover_rate = crossover_rate
        self.mutation_rate = mutation_rate
        self.num_generations = num_generations

    def load_items_from_file(self, filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            capacity = int(next(reader)[0])  # Capacidade
            num_items = int(next(reader)[0])  # Número de itens
            items = []
            for row in reader:
                name, weight, value = row
                items.append(Item(name, int(weight), int(value)))
        return items, capacity

    def run(self):
        population = Population(self.population_size, len(self.items))
        best_fitness_history = []  # Para armazenar a evolução da melhor solução
        best_solution = 0  # Definindo um valor padrão

        for gen in range(self.num_generations):
            population.evolve(self.items, self.capacity, self.crossover_rate, self.mutation_rate)
            # Melhor solução da geração atual
            best_solution = max(population.individuals,
                                key=lambda ind: ind.calculate_fitness(self.items, self.capacity))
            best_fitness_history.append(best_solution.fitness)
        return best_solution, best_fitness_history

    def save_results_to_excel(self, results, filename='../data/results.xlsx'):
        """Salva os resultados dos testes em um arquivo Excel com células centralizadas."""
        headers = ['Teste', 'Crossover', 'Mutação', 'População', 'Gerações', 'Média da Aptidão', 'Melhor Aptidão']

        # Cria um DataFrame a partir dos resultados
        df = pd.DataFrame(results, columns=headers)

        # Salva o DataFrame em um arquivo Excel
        df.to_excel(filename, index=False)

        # Carrega o arquivo Excel salvo
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Centraliza o conteúdo de todas as células
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Salva o arquivo Excel com as modificações
        workbook.save(filename)
