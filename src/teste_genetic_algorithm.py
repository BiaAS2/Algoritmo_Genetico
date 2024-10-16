import matplotlib.pyplot as plt
from knapsack_problem import KnapsackProblem
from genetic_algorithm import GeneticAlgorithm
import numpy as np
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def run_test(problem, population_size, crossover_rate, mutation_rate, num_generations, selection_method, tournament_size, elitism, fitness_metric, num_runs=5):
    """
    Executa múltiplas execuções do algoritmo genético com os parâmetros especificados.

    Args:
        problem (KnapsackProblem): O problema da mochila a ser resolvido.
        population_size (int): O tamanho da população em cada geração.
        crossover_rate (float): A taxa de cruzamento.
        mutation_rate (float): A taxa de mutação.
        num_generations (int): O número de gerações a serem executadas.
        selection_method (str): O método de seleção ('tournament' ou 'roulette').
        tournament_size (int): O tamanho do torneio para o método de seleção por torneio.
        elitism (bool): Se True, usa elitismo.
        fitness_metric (str): A métrica de aptidão a ser usada.
        num_runs (int, optional): O número de execuções do algoritmo. Padrão é 5.

    Returns:
        tuple: Uma tupla contendo a média do melhor fitness e o histórico médio de fitness.
    """
    results = []
    histories = []
    for _ in range(num_runs):
        ga = GeneticAlgorithm(problem, population_size, crossover_rate, mutation_rate, num_generations, selection_method, tournament_size, elitism, fitness_metric)
        _, best_fitness, fitness_history = ga.run()
        results.append(best_fitness)
        histories.append(fitness_history)
    return np.mean(results), np.mean(histories, axis=0)

def test_and_plot(problem, test_params, method_name):
    """
    Executa testes com diferentes parâmetros e plota os resultados.

    Args:
        problem (KnapsackProblem): O problema da mochila a ser resolvido.
        test_params (list): Lista de tuplas, cada uma contendo os parâmetros para um teste.
        method_name (str): Nome do método de seleção sendo testado.

    Returns:
        list: Lista com os resultados coletados.
    """
    num_tests = len(test_params)
    rows = math.ceil(num_tests / 3)
    cols = 3
    results_table = []

    plt.figure(figsize=(15, 5 * rows))

    for i, params in enumerate(test_params, 1):
        population_size, crossover_rate, mutation_rate, num_generations, selection_method, tournament_size, fitness_metric = params
        
        result_with, history_with = run_test(problem, population_size, crossover_rate, mutation_rate, num_generations, selection_method, tournament_size, True, fitness_metric)
        result_without, history_without = run_test(problem, population_size, crossover_rate, mutation_rate, num_generations, selection_method, tournament_size, False, fitness_metric)
        
        plt.subplot(rows, cols, i)
        plt.plot(history_with, label='Com Elitismo')
        plt.plot(history_without, label='Sem Elitismo')
        plt.title(f"Teste {i}: {selection_method.capitalize()}\nPop: {population_size}, CR: {crossover_rate}, MR: {mutation_rate}, Gen: {num_generations}")
        plt.xlabel('Geração')
        plt.ylabel('Fitness')
        plt.legend()

        print(f"Teste {i} ({method_name}):")
        print(f"  Tamanho da População: {population_size}")
        print(f"  Taxa de Cruzamento: {crossover_rate}")
        print(f"  Taxa de Mutação: {mutation_rate}")
        print(f"  Número de Gerações: {num_generations}")
        print(f"  Métrica de Fitness: {fitness_metric}")
        if tournament_size:
            print(f"  Tamanho do Torneio: {tournament_size}")
        print(f"  Fitness Médio com Elitismo: {result_with:.4f}")
        print(f"  Fitness Médio sem Elitismo: {result_without:.4f}")
        print()

        # Armazenar resultados na lista
        results_table.append((i, selection_method, crossover_rate, mutation_rate, population_size, num_generations, result_with, result_without))

    plt.tight_layout()
    plt.show()

    return results_table

def save_results_to_excel(results, filename='Resultados_Torneio_Roleta.xlsx'):
    """Salva os resultados dos testes em um arquivo Excel e centraliza o conteúdo."""
    headers = ['Teste', 'Método', 'Crossover', 'Mutação', 'População', 'Gerações', 'Média com Elitismo', 'Média sem Elitismo']

    # Cria um DataFrame a partir dos resultados
    df = pd.DataFrame(results, columns=headers)

    # Salva o DataFrame em um arquivo Excel sem index
    df.to_excel(filename, index=False)

    # Carrega o workbook criado
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Centraliza o conteúdo de todas as células
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Salva o arquivo Excel com as alterações de estilo
    workbook.save(filename)

def main():
    """
    Função principal que carrega o problema e executa os testes.

    Esta função carrega o problema da mochila a partir de um arquivo,
    define os parâmetros de teste para os métodos de seleção por torneio e roleta,
    e então executa e plota os resultados dos testes.

    Returns:
        None: Esta função não retorna nada, mas gera gráficos e imprime resultados.
    """
    try:
        problem = KnapsackProblem.load_from_file("../data/instancia.csv")
    except Exception as e:
        print(f"Erro ao carregar o problema: {e}")
        return

    # Testes com o método de seleção 'tournament'
    tournament_params = [
        (100, 0.7, 0.01, 100, 'tournament', 3, "maximize_benefit_weight"),
        (200, 0.75, 0.05, 150, 'tournament', 5, "maximize_benefit"),
        (150, 0.8, 0.1, 200, 'tournament', 4, "maximize_benefit_weight"),
        (250, 0.85, 0.03, 120, 'tournament', 6, "maximize_benefit"),
        (180, 0.9, 0.08, 180, 'tournament', 5, "maximize_benefit_weight")
    ]

    print("Resultados para seleção por Torneio:")
    results_tournament = test_and_plot(problem, tournament_params, "Torneio")

    # Testes com o método de seleção 'roulette'
    roulette_params = [
        (100, 0.7, 0.01, 100, 'roulette', None, "maximize_benefit_weight"),
        (200, 0.75, 0.05, 150, 'roulette', None, "maximize_benefit"),
        (150, 0.8, 0.1, 200, 'roulette', None, "maximize_benefit_weight"),
        (250, 0.85, 0.03, 120, 'roulette', None, "maximize_benefit"),
        (180, 0.9, 0.08, 180, 'roulette', None, "maximize_benefit_weight")
    ]

    print("Resultados para seleção por Roleta:")
    results_roulette = test_and_plot(problem, roulette_params, "Roleta")

    # Combine os resultados para salvar
    all_results = results_tournament + results_roulette

    # Salve todos os resultados em um arquivo Excel
    save_results_to_excel(all_results, "Resultados_Torneio_Roleta.xlsx")

if __name__ == "__main__":
    main()
